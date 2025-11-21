import sys
import threading
import requests
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QTextEdit, QPushButton,
    QProgressBar, QTabWidget, QMessageBox, QHBoxLayout, QToolBar, QAction
)
import webbrowser
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont, QPixmap, QCursor
import final_parser2 as vbd
from openpyxl import load_workbook, Workbook
import csv
import os

APP_VERSION = "1.2.0"   # текущая версия программы
GITHUB_REPO = "Oleger1000/SRSPU_Ved_Scrapper"  # твой репозиторий


def resource_path(relative_path):
    """Получаем путь к ресурсам иконок как для dev, так и для PyInstaller"""
    try:
        # PyInstaller создаёт временную папку _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

class VedGUI(QWidget):
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)
    update_available_signal = pyqtSignal(str, str)  # добавляем этот сигнал для версии и URL

    def __init__(self):
        super().__init__()
        self.setWindowTitle("VED Scraper")
        self.resize(600, 570)
        self.update_available_signal.connect(self.show_update_popup)
        self.setFont(QFont("Segoe UI", 10))

        # Начальная тема - светлая
        self.current_theme = 'light'
        self.apply_theme()

        main_layout = QVBoxLayout()
        self.setLayout(main_layout)

        # --- Вкладки ---
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # --- Вкладка 1: Сбор оценок ---
        self.tab_scrape = QWidget()
        self.tabs.addTab(self.tab_scrape, "Сбор оценок")
        self.init_scrape_tab()

        # --- Вкладка 2: Замена ФИО ---
        self.tab_fio = QWidget()
        self.tabs.addTab(self.tab_fio, "Замена ФИО")
        self.init_fio_tab()

        # --- Вкладка 3: О программе ---
        self.tab_about = QWidget()
        self.tabs.addTab(self.tab_about, "О программе")
        self.init_about_tab()
    
        # Проверяем обновления при запуске
        threading.Thread(target=self.check_for_updates, daemon=True).start()


    # ---------- Темы ----------

    def check_for_updates(self):
        import requests

        url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"

        try:
            r = requests.get(url, timeout=5)
            if r.status_code != 200:
                return

            data = r.json()

            raw_name = data.get("name", "").lstrip("v")
            latest_tag = raw_name.strip()
            release_url = data.get("html_url", "")

            def version_tuple(v):
                return tuple(map(int, v.split("."))) if v else (0,)

            if version_tuple(latest_tag) > version_tuple(APP_VERSION):
                # ВАЖНО: уведомляем GUI-поток через сигнал
                self.update_available_signal.emit(latest_tag, release_url)

        except Exception:
            pass

    
    def show_update_popup(self, latest_tag, release_url):

        open_msg = QMessageBox(self)
        open_msg.setWindowTitle("Обновление доступно")
        open_msg.setText(
            f"Вышла новая версия: {latest_tag}\n"
            f"Открыть страницу релизов?"
        )
        open_msg.addButton("Открыть GitHub", QMessageBox.AcceptRole)
        open_msg.addButton("Позже", QMessageBox.RejectRole)

        if open_msg.exec() == QMessageBox.AcceptRole:
            import webbrowser
            webbrowser.open(release_url)




    def update_icons(self):
        self.telegram_label.setPixmap(
            QPixmap(resource_path("assets/telegram.png")).scaled(32, 32, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        )

        github_icon = "assets/github_light.png" if self.current_theme == "dark" else "assets/github.png"
        self.github_label.setPixmap(
            QPixmap(resource_path(github_icon)).scaled(32, 32, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        )

    def apply_theme(self):
        if self.current_theme == 'light':
            self.setStyleSheet("""
                QWidget { background-color: #f3f3f3; color: #000; }
                QGroupBox { border: 1px solid #999; border-radius: 5px; margin-top: 10px; padding: 10px; background-color: #fff; }
                QPushButton { background-color: #0078D7; color: white; border-radius: 5px; padding: 6px; }
                QPushButton:hover { background-color: #005A9E; }
                QTextEdit { border: 1px solid #999; border-radius: 5px; padding: 5px; background-color: #fff; color: #000; }
                QProgressBar { border: 1px solid #999; border-radius: 5px; text-align: center; background-color: #e0e0e0; color: #000; }
                QProgressBar::chunk { background-color: #0078D7; width: 20px; }
                QLabel { color: #000; }
                QToolBar { background: #ddd; border: none; }

                /* вкладки */
                QTabWidget::pane { border: 1px solid #aaa; background: #f3f3f3; }
                QTabBar::tab {
                    background: #e6e6e6;
                    color: #000;
                    padding: 8px 16px;
                    border: 1px solid #aaa;
                    border-bottom: none;
                    border-top-left-radius: 5px;
                    border-top-right-radius: 5px;
                    min-height: 22px;
                    min-width: 110px; 
                    font-size: 15px;
                }
                QTabBar::tab:selected {
                    background: #ffffff;
                    font-weight: bold;
                }
                QTabBar::tab:hover {
                    background: #f9f9f9;
                }
            """)
        else:
            self.setStyleSheet("""
                QWidget { background-color: #2b2b2b; color: #f0f0f0; }
                QGroupBox { border: 1px solid #555; border-radius: 5px; margin-top: 10px; padding: 10px; background-color: #3c3c3c; }
                QPushButton { background-color: #0078D7; color: white; border-radius: 5px; padding: 6px; }
                QPushButton:hover { background-color: #005A9E; }
                QTextEdit { border: 1px solid #555; border-radius: 5px; padding: 5px; background-color: #1e1e1e; color: #f0f0f0; }
                QProgressBar { border: 1px solid #555; border-radius: 5px; text-align: center; background-color: #1e1e1e; color: #f0f0f0; }
                QProgressBar::chunk { background-color: #0078D7; width: 20px; }
                QLabel { color: #f0f0f0; }
                QToolBar { background: #3c3c3c; border: none; }

                /* вкладки */
                QTabWidget::pane { border: 1px solid #444; background: #2b2b2b; }
                QTabBar::tab {
                    background: #3c3c3c;
                    color: #f0f0f0;
                    padding: 8px 16px;
                    border: 1px solid #555;
                    border-bottom: none;
                    border-top-left-radius: 5px;
                    border-top-right-radius: 5px;
                    min-height: 22px;
                    min-width: 110px; 
                    font-size: 15px;
                }
                QTabBar::tab:selected {
                    background: #1e1e1e;
                    font-weight: bold;
                }
                QTabBar::tab:hover {
                    background: #333;
                }
            """)

    def toggle_theme(self):
        self.current_theme = 'dark' if self.current_theme == 'light' else 'light'
        self.apply_theme()
        self.update_icons()
        

    # ---------- Вкладка Сбор оценок ----------
    def init_scrape_tab(self):
        layout = QVBoxLayout()
        self.tab_scrape.setLayout(layout)

        # --- Поля для логина/пароля cookie-сервера ---
        self.api_login_input = QTextEdit()
        self.api_login_input.setPlaceholderText("Email")
        self.api_login_input.setFixedHeight(37)
        layout.addWidget(QLabel("Login:"))
        layout.addWidget(self.api_login_input)

        self.api_password_input = QTextEdit()
        self.api_password_input.setPlaceholderText("Password")
        self.api_password_input.setFixedHeight(37)
        layout.addWidget(QLabel("Password:"))
        layout.addWidget(self.api_password_input)

        self.run_button = QPushButton("Запустить сбор оценок")
        self.run_button.clicked.connect(self.run_scraper_thread)
        layout.addWidget(self.run_button)

        self.progress_bar = QProgressBar()
        self.progress_bar.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.progress_bar)

        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        layout.addWidget(QLabel("Логи:"))
        layout.addWidget(self.log_output)

        # --- ВАЖНО: соединяем сигнал после создания виджетов ---
        self.log_signal.connect(self.log_output.append)
        self.progress_signal.connect(self.progress_bar.setValue)

    # ---------- Вкладка Замена ФИО ----------
    def autofill_rec_numbers(self):
        """Автозаполняем поле соответствий номеров зачеток из XLSX (уникальные значения)."""
        xlsx_path = os.path.join(vbd.OUT_DIR, "all_students.xlsx")

        if not os.path.exists(xlsx_path):
            self.log_signal.emit(f"[!] XLSX с оценками не найден: {xlsx_path}")
            return

        wb = load_workbook(xlsx_path)
        ws = wb.active

        rec_numbers = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            student_id, student_name, discipline, score = row
            if student_name:
                rec_numbers.add(student_name)

        sorted_numbers = sorted(rec_numbers)

        lines = [f"{num}, " for num in sorted_numbers]
        self.fio_input.setPlainText("\n".join(lines))
        self.log_signal.emit(f"[+] Поле заполнено {len(lines)} номерами зачеток (из XLSX)")




    def init_fio_tab(self):
        layout = QVBoxLayout()
        self.tab_fio.setLayout(layout)

        layout.addWidget(QLabel("Введите соответствия номеров зачеток и ФИО по одной паре на строку через запятую:"))
        self.fio_input = QTextEdit()
        self.fio_input.setPlaceholderText("20221001, Иванов Иван Иванович\n20221002, Петров Пётр Петрович")
        layout.addWidget(self.fio_input)

        # --- Кнопка замены ---
        self.replace_button = QPushButton("Заменить на ФИО")
        self.replace_button.clicked.connect(self.replace_ids_with_fio)
        layout.addWidget(self.replace_button)

        # --- Кнопка автозаполнения ---
        autofill_btn = QPushButton("Автозаполнить номера зачеток")
        autofill_btn.clicked.connect(self.autofill_rec_numbers)
        layout.addWidget(autofill_btn)

        # Можно сразу автозаполнить при открытии вкладки:
        self.autofill_rec_numbers()

    # ---------- Вкладка О программе ----------

    def init_about_tab(self):
        layout = QVBoxLayout()
        self.tab_about.setLayout(layout)

        about_text = QTextEdit()
        about_text.setReadOnly(True)
        about_text.setHtml("""
            <h2>VED Scraper</h2>
            <p><b>Версия:</b> 1.2</p>
            <p><b>Автор:</b> @Oleger12</p>
            <p>Программа для сбора оценок студентов с dec.srspu.ru и замены номеров зачеток на ФИО.</p>
            <p>Инструкция:</p>
            <ol>
                <li>Войдите в свой аккаунт на dec.srspu.ru в своем браузере, перезапустите его и программу.</li>
                <li>Перейдите на вкладку "Сбор оценок" и нажмите "Запустить сбор оценок".</li>
                <li>После генерации CSV перейдите на вкладку "Замена ФИО".</li>
                <li>Впишите соответствия номеров зачеток и ФИО по одной паре на строку через запятую.</li>
                <li>Нажмите "Заменить на ФИО в CSV" для создания нового файла.</li>
            </ol>
        """)
        layout.addWidget(about_text)

        # --- Кнопка смены темы ---
        theme_btn = QPushButton("Переключить тему")
        theme_btn.clicked.connect(self.toggle_theme)
        layout.addWidget(theme_btn)

        # --- Иконки ---
        icon_layout = QHBoxLayout()
        icon_layout.addStretch()

        self.telegram_label = QLabel()
        self.github_label = QLabel()

        self.telegram_label.setCursor(QCursor(Qt.PointingHandCursor))
        self.github_label.setCursor(QCursor(Qt.PointingHandCursor))

        self.telegram_label.mousePressEvent = lambda e: webbrowser.open("https://t.me/Oleger12")
        self.github_label.mousePressEvent = lambda e: webbrowser.open("https://github.com/Oleger1000")

        icon_layout.addWidget(self.telegram_label)
        icon_layout.addWidget(self.github_label)
        icon_layout.addStretch()
        layout.addLayout(icon_layout)

        # Инициализация иконок в зависимости от темы
        self.update_icons()

    # ---------- helpers ----------
    def log(self, msg):
        self.log_output.append(msg)
        self.log_output.verticalScrollBar().setValue(
            self.log_output.verticalScrollBar().maximum()
        )

    def init_session_and_fetch_data(self, login: str, password: str):
        """
        Инициализация requests.Session(), получение cookies (локально или с cookie-server),
        тестирование валидности сессии, возврат session или None.
        """
        import requests
        s = requests.Session()

        # 1) Пробуем локальные куки
        cj = vbd.get_cookiejar_for_domain("dec.srspu.ru")
        if cj:
            vbd.inject_cookiejar_into_session(s, cj, "dec.srspu.ru")
            self.log_signal.emit(f"[+] Cookies найдены локально в браузере.")
        else:
            self.log_signal.emit(f"[*] Локальные куки не найдены, запрашиваем с cookie-server...")
            API_URL = "http://89.169.12.12:63592"
            API_KEY = "transfer_train_never_been_located"
            cookies = vbd.fetch_cookies_from_cookie_server(API_URL, API_KEY, login, password)
            if not cookies:
                self.log_signal.emit("[!] Не удалось получить cookies с cookie-server.")
                return None
            vbd.transfer_cookies_from_playwright_format(s, cookies)
            self.log_signal.emit("[+] Cookies подставлены в сессию из cookie-server.")

        # 2) Проверяем валидность сессии
        try:
            test = s.get(vbd.DISCIPLINES_URL, allow_redirects=True)
            if "Login.aspx" in test.url or test.status_code != 200:
                self.log_signal.emit("[!] Сессия невалидна после подстановки cookies.")
                return None
        except Exception as e:
            self.log_signal.emit(f"[!] Ошибка при проверке сессии: {e}")
            return None

        return s

    def scrape_discipline_data(self, session: requests.Session):
        """
        Основная логика сбора студентов и оценок, сохранение HTML и CSV.
        """
        r = session.get(vbd.DISCIPLINES_URL)
        r.raise_for_status()
        discipline_id, discipline_name, discipline_url = vbd.get_first_available_discipline(r.text)
        if not discipline_id:
            self.log_signal.emit("[!] Нет доступных дисциплин.")
            return

        self.log_signal.emit(f"[*] Первая дисциплина: {discipline_name} (id={discipline_id})")
        r_disc = session.get(discipline_url)
        r_disc.raise_for_status()
        group_id, group_name = vbd.extract_group_from_discipline_page(r_disc.text)
        if not group_id:
            self.log_signal.emit("[!] Не удалось получить group_id.")
            return
        self.log_signal.emit(f"[*] Группа: {group_name} (id={group_id})")

        group_html = vbd.fetch_group_page(session, group_id)
        students = vbd.extract_student_ids_and_names(group_html)
        total_students = len(students)
        self.progress_signal.emit(0)
        self.progress_bar.setMaximum(total_students)

        rows_for_csv = []
        for idx, (sid, student_name) in enumerate(students.items(), start=1):
            try:
                ved_html = vbd.fetch_totalved_for_student(session, sid)
            except Exception as e:
                self.log_signal.emit(f"    !! Ошибка при загрузке {sid}: {e}")
                continue

            html_file = os.path.join(vbd.HTML_DIR, f"student_{sid}.html")
            with open(html_file, "w", encoding="utf-8") as f:
                f.write(ved_html)
            self.log_signal.emit(f"    -> сохранено {html_file}")

            discipline_scores = vbd.parse_totalved_discipline_scores(ved_html)
            if discipline_scores:
                for discipline, score in discipline_scores:
                    rows_for_csv.append([sid, student_name, discipline, score])
            else:
                rows_for_csv.append([sid, student_name, "NO_DISCIPLINES", ""])

            self.progress_signal.emit(idx)
            vbd.time.sleep(vbd.REQUESTS_SLEEP)

        if rows_for_csv:
            header = ["student_id", "student_name", "discipline", "score"]
            xlsx_path = os.path.join(vbd.OUT_DIR, "all_students.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.append(["student_id", "student_name", "discipline", "score"])

            for row in rows_for_csv:
                ws.append(row)

            wb.save(xlsx_path)
            self.log_signal.emit(f"[+] XLSX сохранён: {xlsx_path}")

        else:
            self.log_signal.emit("[!] Нечего сохранять в CSV.")

    # ---------- Парсер ----------
    def run_scraper_thread(self):
        self.run_button.setEnabled(False)
        threading.Thread(target=self.run_scraper, daemon=True).start()

    def run_scraper(self):
        import builtins
        original_print = builtins.print

        def gui_print(*args, **kwargs):
            msg = " ".join(str(a) for a in args)
            self.log_signal.emit(msg)
            original_print(*args, **kwargs)

        builtins.print = gui_print

        try:
            self.run_button.setEnabled(False)
            self.log_signal.emit("[*] Инициализация сессии...")

            # --- Получаем login/password из GUI ---
            login = self.api_login_input.toPlainText().strip()
            password = self.api_password_input.toPlainText().strip()
            if not login or not password:
                self.log_signal.emit("[!] Укажите login и password для cookie-сервера")
                return

            # --- Инициализация сессии и получение cookies ---
            session = self.init_session_and_fetch_data(login, password)
            if not session:
                self.log_signal.emit("[!] Сессия невалидна. Прекращаем выполнение.")
                return

            # --- Сбор данных ---
            self.scrape_discipline_data(session)
            self.log_signal.emit("[*] Готово.")

        except Exception as e:
            self.log_signal.emit(f"[!] Ошибка: {e}")

        finally:
            builtins.print = original_print
            self.run_button.setEnabled(True)

    # ---------- Заменяем номера зачеток на ФИО ----------
    def replace_ids_with_fio(self):
        text = self.fio_input.toPlainText().strip()
        if not text:
            QMessageBox.warning(self, "Ошибка", "Введите соответствия номеров зачеток и ФИО")
            return

        mapping = {}
        for line in text.splitlines():
            parts = line.split(",", 1)
            if len(parts) != 2:
                self.log(f"[!] Неправильный формат строки: {line}")
                continue
            rec_number, full_name = parts[0].strip(), parts[1].strip()
            mapping[rec_number] = full_name

        xlsx_path = os.path.join(vbd.OUT_DIR, "all_students.xlsx")
        if not os.path.exists(xlsx_path):
            QMessageBox.warning(self, "Ошибка", f"XLSX-файл с оценками не найден: {xlsx_path}")
            return

        wb = load_workbook(xlsx_path)
        ws = wb.active

        not_found = set()

        # Делаем копию как новый файл
        new_xlsx = os.path.join(vbd.OUT_DIR, "all_students_with_names.xlsx")
        new_wb = Workbook()
        new_ws = new_wb.active

        # Переносим заголовок
        header = ["student_id", "student_name", "discipline", "score"]
        new_ws.append(header)

        for row in ws.iter_rows(min_row=2, values_only=True):
            student_id, student_name, discipline, score = row
            if student_name in mapping:
                student_name = mapping[student_name]
            else:
                not_found.add(student_name)

            new_ws.append([student_id, student_name, discipline, score])

        new_wb.save(new_xlsx)

        self.log(f"[+] Замена выполнена. Новый XLSX сохранён: {new_xlsx}")

        if not_found:
            self.log(f"[!] Не найдены соответствия для: {', '.join(not_found)}")

        QMessageBox.information(self, "Готово", f"Новый XLSX сохранён:\n{new_xlsx}")
    

if __name__ == "__main__":
    app = QApplication(sys.argv)
    gui = VedGUI()
    gui.show()
    sys.exit(app.exec_())
